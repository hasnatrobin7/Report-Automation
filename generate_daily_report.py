import pandas as pd
import numpy as np
import os
import concurrent.futures
import duckdb
from datetime import datetime, time, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
import glob
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import win32com.client
import tempfile
import base64
from PIL import Image
import io

# Constants
SYNDROM_DB = 'SyndromDB'
REPORT_FILE = 'Daily_TLA_Report.xlsx'
EXCLUDE_FILE = 'exclude_syndroms.txt'
RECIPIENTS_FILE = 'recipients.txt'
IMG_WIDTH = 80
IMG_HEIGHT = 60
# NEW: Only load the columns actually needed from the raw Excel files
ESSENTIAL_COLS = ['StartDateTime', 'Syndrom', 'SyndromStatus', 'UUT', 'SerialNumber']
# Folder where per-workbook parquet caches are stored
PARQUET_CACHE_DIR = "_parquet_cache"

# Ensure cache directory exists
os.makedirs(PARQUET_CACHE_DIR, exist_ok=True)

# Shift time boundaries
SHIFT_1_START = time(0, 0)
SHIFT_1_END = time(15, 30)
SHIFT_2_START = time(15, 30)
SHIFT_2_END = time(23, 59, 59)

def get_shift(dt):
    """Return the shift label (1st/2nd) for a pandas.Timestamp or datetime."""
    t = dt.time()
    if SHIFT_1_START <= t < SHIFT_1_END:
        return '1st Shift'
    elif SHIFT_2_START <= t <= SHIFT_2_END:
        return '2nd Shift'
        return 'Unknown'

# Vectorised version for DataFrames (avoids Python-level loops)
def vectorized_shift(series):
    """Vectorised helper to map a datetime series into shift labels."""
    times = series.dt.time
    return pd.Series(
        np.where(
            (times >= SHIFT_1_START) & (times < SHIFT_1_END), '1st Shift',
            np.where((times >= SHIFT_2_START) & (times <= SHIFT_2_END), '2nd Shift', 'Unknown')
        ),
        index=series.index
    )

def load_exclude_list():
    if not os.path.exists(EXCLUDE_FILE):
        return set()
    with open(EXCLUDE_FILE, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    # Remove comments and whitespace
    return set(line.strip() for line in lines if line.strip() and not line.strip().startswith('#'))

def merge_consecutive_cells(ws, col_idx):
    """Merge consecutive cells in a column if they have the same value (except header)."""
    col_letter = get_column_letter(col_idx + 1)
    start_row = 2  # skip header
    end_row = ws.max_row
    prev_val = None
    merge_start = start_row
    for row in range(start_row, end_row + 1):
        val = ws[f'{col_letter}{row}'].value
        if val != prev_val:
            if row - merge_start > 1 and prev_val is not None:
                ws.merge_cells(f'{col_letter}{merge_start}:{col_letter}{row-1}')
            merge_start = row
            prev_val = val
    # Merge last group
    if end_row - merge_start >= 1 and prev_val is not None:
        ws.merge_cells(f'{col_letter}{merge_start}:{col_letter}{end_row}')

def sanitize_syndrom_name(syndrom):
    """Convert syndrom name to folder-safe name by replacing special characters."""
    # Replace special characters that can't be used in Windows folder names
    # / \ : * ? " < > | are not allowed in Windows folder names
    replacements = {
        '/': '-',
        '\\': '-', 
        ':': '-',
        '*': '-',
        '?': '-',
        '"': '-',
        '<': '-',
        '>': '-',
        '|': '-',
        ',': ',',  # Keep commas as they appear in existing folder names
        ' ': ' '   # Keep spaces
    }
    
    sanitized = syndrom
    for char, replacement in replacements.items():
        sanitized = sanitized.replace(char, replacement)
    
    return sanitized

def get_syndrom_db_info(syndrom):
    """Return (golden_img_path, defect_img_path, description) for a syndrom, or (None, None, None) if not found."""
    # First try exact match
    folder = os.path.join(SYNDROM_DB, syndrom)
    if os.path.exists(folder):
        golden_img = os.path.join(folder, 'golden.jpg')
        defect_img = os.path.join(folder, 'defect.jpg')
        desc_file = os.path.join(folder, 'description.txt')
        golden_img = golden_img if os.path.exists(golden_img) else None
        defect_img = defect_img if os.path.exists(defect_img) else None
        description = None
        if os.path.exists(desc_file):
            with open(desc_file, 'r', encoding='utf-8') as f:
                description = f.read().strip()
        return golden_img, defect_img, description
    
    # If exact match fails, try sanitized name
    sanitized_syndrom = sanitize_syndrom_name(syndrom)
    folder = os.path.join(SYNDROM_DB, sanitized_syndrom)
    golden_img = os.path.join(folder, 'golden.jpg')
    defect_img = os.path.join(folder, 'defect.jpg')
    desc_file = os.path.join(folder, 'description.txt')
    golden_img = golden_img if os.path.exists(golden_img) else None
    defect_img = defect_img if os.path.exists(defect_img) else None
    description = None
    if os.path.exists(desc_file):
        with open(desc_file, 'r', encoding='utf-8') as f:
            description = f.read().strip()
    return golden_img, defect_img, description

def find_excel_files():
    """Find all Excel files in the current directory and extract their dates."""
    # Collect *.xlsx files except the report workbook and any temporaries
    excel_files = [f for f in glob.glob("*.xlsx") if f not in {REPORT_FILE}]
    file_dates = []
    
    for file in excel_files:
        try:
            # Read ONLY the StartDateTime column for min/max extraction
            df = pd.read_excel(file, usecols=['StartDateTime'])
            if not df.empty:
                df['StartDateTime'] = pd.to_datetime(df['StartDateTime'])
                min_date = df['StartDateTime'].min().date()
                max_date = df['StartDateTime'].max().date()
                file_dates.append({
                    'file': file,
                    'min_date': min_date,
                    'max_date': max_date,
                    'date_range': f"{min_date} to {max_date}"
                })
        except Exception as e:
            print(f"Warning: Could not read {file}: {e}")
    
    return file_dates

def get_user_date_selection(file_dates, purpose="report"):
    """Interactive date selection."""
    if not file_dates:
        print("No Excel files found with date data!")
        return None, None
    
    print(f"\nAvailable date ranges for {purpose}:")
    for i, file_info in enumerate(file_dates, 1):
        print(f"{i}. {file_info['file']} - {file_info['date_range']}")
    
    print("\nOptions:")
    print("1. Single date")
    print("2. Date range")
    print("3. Latest date")
    print("4. All available data")
    
    while True:
        try:
            choice = input("\nEnter your choice (1-4): ").strip()
            if choice == '1':
                # Single date
                print("\nAvailable dates:")
                all_dates = set()
                for file_info in file_dates:
                    all_dates.add(file_info['min_date'])
                    all_dates.add(file_info['max_date'])
                
                sorted_dates = sorted(all_dates)
                for i, date in enumerate(sorted_dates, 1):
                    print(f"{i}. {date}")
                
                date_choice = int(input("Enter date number: ")) - 1
                selected_date = sorted_dates[date_choice]
                return selected_date, selected_date
                
            elif choice == '2':
                # Date range
                print("\nEnter start date (YYYY-MM-DD):")
                start_date = input("Start date: ").strip()
                end_date = input("End date: ").strip()
                return datetime.strptime(start_date, '%Y-%m-%d').date(), datetime.strptime(end_date, '%Y-%m-%d').date()
                
            elif choice == '3':
                # Latest date
                latest_date = max(file_info['max_date'] for file_info in file_dates)
                print(f"Using latest date: {latest_date}")
                return latest_date, latest_date
                
            elif choice == '4':
                # All available data
                min_date = min(file_info['min_date'] for file_info in file_dates)
                max_date = max(file_info['max_date'] for file_info in file_dates)
                print(f"Using all data from {min_date} to {max_date}")
                return min_date, max_date
                
            else:
                print("Invalid choice. Please enter 1-4.")
        except (ValueError, IndexError) as e:
            print(f"Invalid input: {e}. Please try again.")

# ------------------------------------------------------------------
# Helper for parallel reading
def _read_and_filter(args):
    """Read an Excel file and return only rows within the date window.

    This is a top-level function so it can be pickled for multiprocessing.
    Parameters
    ----------
    args : tuple(file_path, start_date, end_date)
    """
    file, start_date, end_date = args
    try:
        import_path = file
        parquet_path = os.path.join(PARQUET_CACHE_DIR, os.path.basename(file) + ".parquet")

        # Load from parquet cache if up-to-date
        if os.path.exists(parquet_path) and os.path.getmtime(parquet_path) >= os.path.getmtime(import_path):
            df = pd.read_parquet(parquet_path, columns=ESSENTIAL_COLS)
        else:
            df = pd.read_excel(import_path, usecols=lambda c: c in ESSENTIAL_COLS, parse_dates=['StartDateTime'])
            # Save to cache for next run
            try:
                df.to_parquet(parquet_path, index=False)
            except Exception:
                # Skip caching if pyarrow not available
                pass

        if 'StartDateTime' not in df.columns or df.empty:
            return None

        mask = (df['StartDateTime'].dt.date >= start_date) & (df['StartDateTime'].dt.date <= end_date)
        filtered_df = df.loc[mask]
        if filtered_df.empty:
            return None
        return filtered_df
    except Exception as e:
        return f"ERROR::{file}::{e}"

# ------------------------------------------------------------------
# DuckDB integration: cache management + fast filtered loading

def _convert_excel_to_parquet(excel_file, parquet_path):
    """Convert one Excel workbook to Parquet containing only ESSENTIAL_COLS."""
    df = pd.read_excel(excel_file, usecols=lambda c: c in ESSENTIAL_COLS, parse_dates=['StartDateTime'])
    if df.empty:
        return False
    df.to_parquet(parquet_path, index=False)
    return True


def ensure_parquet_cache(excel_files):
    """Ensure every XLSX in the list has an up-to-date Parquet cache."""
    outdated = []
    for f in excel_files:
        parquet_path = os.path.join(PARQUET_CACHE_DIR, os.path.basename(f) + ".parquet")
        if not os.path.exists(parquet_path) or os.path.getmtime(parquet_path) < os.path.getmtime(f):
            outdated.append((f, parquet_path))

    if not outdated:
        return

    # Convert sequentially; Excel parsing is heavy but parallel gains are limited by GIL anyway.
    # We keep it simple; you can switch to ProcessPool if desired.
    for src, dst in outdated:
        try:
            print(f"Caching {os.path.basename(src)} → parquet …", end=" ")
            _convert_excel_to_parquet(src, dst)
            print("done")
        except Exception as e:
            print(f"failed ({e})")


def load_data_duckdb(start_date, end_date, excel_files):
    """Load filtered data using DuckDB over Parquet caches."""
    ensure_parquet_cache(excel_files)

    start_ts = pd.Timestamp(start_date)
    # Use inclusive <= end
    end_ts = pd.Timestamp(end_date) + pd.Timedelta(days=1) - pd.Timedelta(microseconds=1)

    parquet_glob = os.path.join(PARQUET_CACHE_DIR, "*.parquet")

    con = duckdb.connect()
    query = (
        f"SELECT * FROM parquet_scan('{parquet_glob}') "
        f"WHERE StartDateTime >= '{start_ts}' AND StartDateTime <= '{end_ts}'"
    )
    df = con.execute(query).df()
    if df.empty:
        return None
    return df

def load_data_for_date_range(start_date, end_date, files_in_range=None):
    """Load and combine data from Excel files within the given date range.

    Parameters
    ----------
    start_date, end_date : datetime.date
        The inclusive date window.
    files_in_range : list[str] or None
        If provided, only these files will be scanned (much faster when caller pre-filters).
    """
    all_data = []

    # Decide which files to inspect
    excel_files = files_in_range if files_in_range is not None else [f for f in glob.glob("*.xlsx") if f not in {REPORT_FILE}]

    # Read files in parallel for speed (I/O + CPU heavy Excel parsing)
    # NOTE: On Windows we need to be inside the '__main__' guard (which we are)
    if len(excel_files) > 1:
        tasks = [(f, start_date, end_date) for f in excel_files]
        with concurrent.futures.ProcessPoolExecutor(max_workers=min(4, os.cpu_count() or 1)) as pool:
            for result in pool.map(_read_and_filter, tasks):
                if result is None:
                    continue
                if isinstance(result, str) and result.startswith("ERROR::"):
                    _, file, err = result.split("::", 2)
                    print(f"Warning: Could not read {file}: {err}")
                    continue
                all_data.append(result)
    else:
        # Single file: fall back to in-process read (avoids process overhead)
        for file in excel_files:
            try:
                parquet_path = os.path.join(PARQUET_CACHE_DIR, os.path.basename(file) + ".parquet")
                if os.path.exists(parquet_path) and os.path.getmtime(parquet_path) >= os.path.getmtime(file):
                    df = pd.read_parquet(parquet_path, columns=ESSENTIAL_COLS)
                else:
                    df = pd.read_excel(file, usecols=lambda c: c in ESSENTIAL_COLS, parse_dates=['StartDateTime'])
                    try:
                        df.to_parquet(parquet_path, index=False)
                    except Exception:
                        pass
                if 'StartDateTime' not in df.columns:
                    continue
                mask = (df['StartDateTime'].dt.date >= start_date) & (df['StartDateTime'].dt.date <= end_date)
                filtered_df = df.loc[mask]
                if not filtered_df.empty:
                    all_data.append(filtered_df)
                    print(f"Loaded {len(filtered_df)} records from {file}")
            except Exception as e:
                print(f"Warning: Could not read {file}: {e}")
    
    if not all_data:
        print("No data found for the selected date range!")
        return None
    
    # Combine all data
    combined_df = pd.concat(all_data, ignore_index=True)
    print(f"Total records loaded: {len(combined_df)}")
    return combined_df

def calculate_trend_data(df, top_syndroms):
    """Calculate daily and weekly trend data for the specified top syndroms."""
    # Add date column
    df['Date'] = df['StartDateTime'].dt.date
    
    # Calculate daily trends for the specified top syndroms
    daily_trends = []
    for syndrom in top_syndroms:
        syndrom_fails = df[df['Syndrom'] == syndrom]
        daily_fails = syndrom_fails.groupby('Date').size()
        daily_total = df.groupby('Date').size()
        daily_rate = (daily_fails / daily_total * 100).fillna(0)
        
        for date, rate in daily_rate.items():
            daily_trends.append({
                'Date': date,
                'Syndrom': syndrom,
                'Daily_Rate': round(rate, 2)
            })
    
    daily_df = pd.DataFrame(daily_trends)
    
    # Calculate weekly trends for the specified top syndroms
    df['Week'] = df['StartDateTime'].dt.to_period('W')
    weekly_trends = []
    for syndrom in top_syndroms:
        syndrom_fails = df[df['Syndrom'] == syndrom]
        weekly_fails = syndrom_fails.groupby('Week').size()
        weekly_total = df.groupby('Week').size()
        weekly_rate = (weekly_fails / weekly_total * 100).fillna(0)
        
        for week, rate in weekly_rate.items():
            weekly_trends.append({
                'Week': week,
                'Syndrom': syndrom,
                'Weekly_Rate': round(rate, 2)
            })
    
    weekly_df = pd.DataFrame(weekly_trends)
    
    return daily_df, weekly_df

def create_trend_charts(wb, daily_df, weekly_df, top_syndroms):
    """Create trend charts and add to Excel workbook."""
    from openpyxl.chart.axis import ChartLines
    # Pivot daily data: rows=Date, columns=Syndrom, values=Daily_Rate
    daily_pivot = daily_df.pivot(index='Date', columns='Syndrom', values='Daily_Rate').fillna(0)
    daily_ws = wb.create_sheet("Daily Trend")
    # Write header
    daily_ws.append(["Date"] + list(daily_pivot.columns))
    # Write data (convert dates to strings for Excel category axis)
    for date, row in daily_pivot.iterrows():
        daily_ws.append([str(date)] + [row.get(s, 0) for s in daily_pivot.columns])
    # Create LineChart for daily trend
    daily_chart = LineChart()
    daily_chart.title = "Daily Fail Rate Trends"
    daily_chart.x_axis.title = "Date"
    daily_chart.y_axis.title = "Fail Rate (%)"
    daily_chart.height = 15
    daily_chart.width = 20
    data = Reference(daily_ws, min_col=2, min_row=1, max_col=1+len(daily_pivot.columns), max_row=1+len(daily_pivot))
    cats = Reference(daily_ws, min_col=1, min_row=2, max_row=1+len(daily_pivot))
    daily_chart.add_data(data, titles_from_data=True)
    daily_chart.set_categories(cats)
    # Explicitly enable axes and tick labels
    daily_chart.x_axis.majorTickMark = "in"
    daily_chart.y_axis.majorTickMark = "in"
    daily_chart.x_axis.tickLblPos = "nextTo"
    daily_chart.y_axis.tickLblPos = "nextTo"
    daily_chart.x_axis.crosses = "autoZero"
    daily_chart.y_axis.crosses = "autoZero"
    daily_chart.x_axis.majorGridlines = ChartLines()
    daily_chart.y_axis.majorGridlines = ChartLines()
    # Enable data labels for each series
    for ser in daily_chart.series:
        ser.dLbls = None  # Remove data labels for clarity, or set to DataLabelList() to enable
    daily_ws.add_chart(daily_chart, f"{get_column_letter(2+len(daily_pivot.columns))}2")

    # Pivot weekly data: rows=Week, columns=Syndrom, values=Weekly_Rate
    weekly_pivot = weekly_df.pivot(index='Week', columns='Syndrom', values='Weekly_Rate').fillna(0)
    weekly_ws = wb.create_sheet("Weekly Trend")
    # Write header
    weekly_ws.append(["Week"] + list(weekly_pivot.columns))
    # Write data (convert weeks to strings for Excel category axis)
    for week, row in weekly_pivot.iterrows():
        weekly_ws.append([str(week)] + [row.get(s, 0) for s in weekly_pivot.columns])
    # Create LineChart for weekly trend
    weekly_chart = LineChart()
    weekly_chart.title = "Weekly Fail Rate Trends"
    weekly_chart.x_axis.title = "Week"
    weekly_chart.y_axis.title = "Fail Rate (%)"
    weekly_chart.height = 15
    weekly_chart.width = 20
    data = Reference(weekly_ws, min_col=2, min_row=1, max_col=1+len(weekly_pivot.columns), max_row=1+len(weekly_pivot))
    cats = Reference(weekly_ws, min_col=1, min_row=2, max_row=1+len(weekly_pivot))
    weekly_chart.add_data(data, titles_from_data=True)
    weekly_chart.set_categories(cats)
    # Explicitly enable axes and tick labels
    weekly_chart.x_axis.majorTickMark = "in"
    weekly_chart.y_axis.majorTickMark = "in"
    weekly_chart.x_axis.tickLblPos = "nextTo"
    weekly_chart.y_axis.tickLblPos = "nextTo"
    weekly_chart.x_axis.crosses = "autoZero"
    weekly_chart.y_axis.crosses = "autoZero"
    weekly_chart.x_axis.majorGridlines = ChartLines()
    weekly_chart.y_axis.majorGridlines = ChartLines()
    # Enable data labels for each series
    for ser in weekly_chart.series:
        ser.dLbls = None  # Remove data labels for clarity, or set to DataLabelList() to enable
    weekly_ws.add_chart(weekly_chart, f"{get_column_letter(2+len(weekly_pivot.columns))}2")

def create_merged_image_and_description_cells(ws, report_rows, syndrom_col, golden_img_col, defect_img_col, desc_col):
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    syndrom_to_rows = {}
    for idx, row in enumerate(report_rows, start=2):  # start=2 to skip header
        syndrom = row['Monitor Name']
        if syndrom not in syndrom_to_rows:
            syndrom_to_rows[syndrom] = []
        syndrom_to_rows[syndrom].append(idx)
    for syndrom, rows in syndrom_to_rows.items():
        first_row = rows[0]
        last_row = rows[-1]
        # Merge golden image, defect image, and description columns
        if last_row > first_row:
            ws.merge_cells(f'{get_column_letter(golden_img_col)}{first_row}:{get_column_letter(golden_img_col)}{last_row}')
            ws.merge_cells(f'{get_column_letter(defect_img_col)}{first_row}:{get_column_letter(defect_img_col)}{last_row}')
            ws.merge_cells(f'{get_column_letter(desc_col)}{first_row}:{get_column_letter(desc_col)}{last_row}')
        # Insert image only in the first row
        golden_img = report_rows[first_row-2]['Golden Image']
        defect_img = report_rows[first_row-2]['Defect Image']
        description = report_rows[first_row-2]['Description']
        if golden_img:
            img = XLImage(golden_img)
            img.width = 80
            img.height = 60
            ws.add_image(img, f'{get_column_letter(golden_img_col)}{first_row}')
        if defect_img:
            img = XLImage(defect_img)
            img.width = 80
            img.height = 60
            ws.add_image(img, f'{get_column_letter(defect_img_col)}{first_row}')
        # Only set description in the first cell
        ws[f'{get_column_letter(desc_col)}{first_row}'].value = description
        # Clear text in the first image cells (do not assign to merged cells)
        ws[f'{get_column_letter(golden_img_col)}{first_row}'].value = None
        ws[f'{get_column_letter(defect_img_col)}{first_row}'].value = None

def load_recipients():
    """Load email recipients from recipients.txt file."""
    if not os.path.exists(RECIPIENTS_FILE):
        print(f"Warning: {RECIPIENTS_FILE} not found. No email will be sent.")
        return []
    
    with open(RECIPIENTS_FILE, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    # Remove comments and whitespace, filter valid email addresses
    recipients = []
    for line in lines:
        line = line.strip()
        if line and not line.startswith('#') and '@' in line:
            recipients.append(line)
    
    return recipients

def generate_chart_images(daily_df, weekly_df, top_syndroms, start_date, end_date):
    """Generate trend charts as PNG images for email embedding."""
    chart_files = []
    
    # Set up matplotlib style
    plt.style.use('default')
    plt.rcParams['figure.figsize'] = (12, 8)
    plt.rcParams['font.size'] = 10
    
    # Daily trend chart
    if not daily_df.empty:
        plt.figure(figsize=(12, 8))
        for syndrom in top_syndroms:
            syndrom_data = daily_df[daily_df['Syndrom'] == syndrom]
            if not syndrom_data.empty:
                plt.plot(syndrom_data['Date'], syndrom_data['Daily_Rate'], 
                        marker='o', linewidth=2, label=syndrom)
        
        plt.title('Daily Fail Rate Trends', fontsize=14, fontweight='bold')
        plt.xlabel('Date', fontsize=12)
        plt.ylabel('Fail Rate (%)', fontsize=12)
        plt.legend(fontsize=10)
        plt.grid(True, alpha=0.3)
        plt.xticks(rotation=45)
        plt.tight_layout()
        
        daily_chart_file = 'daily_trend_chart.png'
        plt.savefig(daily_chart_file, dpi=300, bbox_inches='tight')
        plt.close()
        chart_files.append(daily_chart_file)
    
    # Weekly trend chart
    if not weekly_df.empty:
        plt.figure(figsize=(12, 8))
        for syndrom in top_syndroms:
            syndrom_data = weekly_df[weekly_df['Syndrom'] == syndrom]
            if not syndrom_data.empty:
                plt.plot(syndrom_data['Week'].astype(str), syndrom_data['Weekly_Rate'], 
                        marker='s', linewidth=2, label=syndrom)
        
        plt.title('Weekly Fail Rate Trends', fontsize=14, fontweight='bold')
        plt.xlabel('Week', fontsize=12)
        plt.ylabel('Fail Rate (%)', fontsize=12)
        plt.legend(fontsize=10)
        plt.grid(True, alpha=0.3)
        plt.xticks(rotation=45)
        plt.tight_layout()
        
        weekly_chart_file = 'weekly_trend_chart.png'
        plt.savefig(weekly_chart_file, dpi=300, bbox_inches='tight')
        plt.close()
        chart_files.append(weekly_chart_file)
    
    return chart_files

def create_email_summary_table(report_rows):
    """Create a summary table for email (pivoted by shift, no SNs)."""
    # Build a dict keyed by (syndrom, UUT)
    summary_data = {}
    for row in report_rows:
        syndrom = row['Monitor Name']
        uut = row['UUT']
        key = (syndrom, uut)
        shift = row['Shift']
        rate = row['Rate']
        if key not in summary_data:
            summary_data[key] = {
                'Monitor Name': syndrom,
                'UUT': uut,
                '1st Shift': '',
                '2nd Shift': '',
                'Golden Image': row['Golden Image'],
                'Defect Image': row['Defect Image'],
                'Description': row['Description']
            }
        if shift == '1st Shift':
            summary_data[key]['1st Shift'] = rate
        elif shift == '2nd Shift':
            summary_data[key]['2nd Shift'] = rate
    # Convert to DataFrame
    summary_df = pd.DataFrame(list(summary_data.values()))
    return summary_df

def create_html_table(df):
    """Convert DataFrame to HTML table for email, embedding images if available."""
    html = '<table border="1" style="border-collapse: collapse; width: 100%; font-family: Arial, sans-serif;">'
    
    # Header
    html += '<tr style="background-color: #f2f2f2; font-weight: bold;">'
    for col in df.columns:
        html += f'<th style="padding: 8px; text-align: left; border: 1px solid #ddd;">{col}</th>'
    html += '</tr>'
    
    # Data rows
    for _, row in df.iterrows():
        html += '<tr>'
        for col in df.columns:
            if col in ['Golden Image', 'Defect Image']:
                img_path = row[col]
                if isinstance(img_path, str) and os.path.exists(img_path):
                    try:
                        # Open and resize image to exactly 400x400 pixels
                        with Image.open(img_path) as img:
                            # Convert to RGB if necessary
                            if img.mode != 'RGB':
                                img = img.convert('RGB')
                            # Resize to exactly 400x400 pixels
                            img_resized = img.resize((400, 400), Image.Resampling.LANCZOS)
                            # Save to bytes
                            img_buffer = io.BytesIO()
                            img_resized.save(img_buffer, format='JPEG', quality=85)
                            img_buffer.seek(0)
                            img_data = base64.b64encode(img_buffer.getvalue()).decode('utf-8')
                        html += f'<td style="padding: 8px; border: 1px solid #ddd; width:400px; height:400px; text-align:center;"><img src="data:image/jpeg;base64,{img_data}" style="width:400px; height:400px;" /></td>'
                    except Exception as e:
                        print(f"Error processing image {img_path}: {e}")
                        html += '<td style="padding: 8px; border: 1px solid #ddd; width:400px; height:400px;"></td>'
                else:
                    html += '<td style="padding: 8px; border: 1px solid #ddd; width:400px; height:400px;"></td>'
            else:
                value = str(row[col]) if pd.notna(row[col]) else ''
                html += f'<td style="padding: 8px; border: 1px solid #ddd;">{value}</td>'
        html += '</tr>'
    
    html += '</table>'
    return html

def send_email_with_charts(recipients, chart_files, html_table, start_date, end_date):
    """Send Outlook email with embedded charts and table."""
    try:
        # Create Outlook application object
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)  # 0 = olMailItem
        
        # Set email properties
        mail.Subject = f"Daily TLA Report - {start_date} to {end_date}"
        mail.To = "; ".join(recipients)
        
        # Create HTML body
        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; margin: 20px;">
            <h2 style="color: #333;">Daily TLA Report</h2>
            <p><strong>Date Range:</strong> {start_date} to {end_date}</p>
            
            <h3 style="color: #555; margin-top: 30px;">Top 3 Syndroms Summary</h3>
            {html_table}
            
            <h3 style="color: #555; margin-top: 30px;">Trend Charts</h3>
        """
        
        # Add chart images
        for chart_file in chart_files:
            if os.path.exists(chart_file):
                html_body += f'<p><img src="{chart_file}" style="max-width: 100%; height: auto;" /></p>'
        
        html_body += """
            <p style="margin-top: 30px; color: #666; font-size: 12px;">
                This report was automatically generated by the Daily TLA Report Generator.
            </p>
        </body>
        </html>
        """
        
        mail.HTMLBody = html_body
        
        # Attach chart images
        for chart_file in chart_files:
            if os.path.exists(chart_file):
                mail.Attachments.Add(os.path.abspath(chart_file))
        
        # Send the email
        mail.Send()
        print(f"Email sent successfully to: {', '.join(recipients)}")
        
        # Clean up chart files
        for chart_file in chart_files:
            if os.path.exists(chart_file):
                os.remove(chart_file)
                
    except Exception as e:
        print(f"Error sending email: {e}")
        # Clean up chart files even if email fails
        for chart_file in chart_files:
            if os.path.exists(chart_file):
                os.remove(chart_file)

def main():
    print("=== Daily TLA Report Generator ===")
    
    # Find available Excel files and dates
    file_dates = find_excel_files()
    if not file_dates:
        print("No Excel files found in the current directory!")
        return
    
    # Get user date selection for main report
    start_date, end_date = get_user_date_selection(file_dates, "main report")
    if start_date is None:
        return
    
    # Pre-filter the file list so we only open spreadsheets that can possibly contain the requested dates
    files_in_range = [info['file'] for info in file_dates if not (info['max_date'] < start_date or info['min_date'] > end_date)]

    # Load data for the selected date range – DuckDB over cached parquet
    df = load_data_duckdb(start_date, end_date, files_in_range)
    if df is None:
        return
    
    # Add shift information (vectorised for speed)
    df['Shift'] = vectorized_shift(df['StartDateTime'])
    
    # Only consider failed tests
    fail_df = df[df['SyndromStatus'].str.lower() != 'pass']
    
    # Exclude syndroms from the list
    exclude_set = load_exclude_list()
    fail_df = fail_df[~fail_df['Syndrom'].isin(exclude_set)]
    
    # Find top 3 syndroms by total fail count for the main report date range
    top_syndroms = (
        fail_df.groupby('Syndrom').size().sort_values(ascending=False).head(3).index.tolist()
    )

    # NEW: Preview the top 3 syndroms to the user
    print("\nTop 3 Syndroms for the selected date range:")
    for idx, syndrom in enumerate(top_syndroms, start=1):
        syndrom_fail_count = fail_df[fail_df['Syndrom'] == syndrom].shape[0]
        print(f"{idx}. {syndrom} - {syndrom_fail_count} fails")
    print("-" * 40)

    top_fail_df = fail_df[fail_df['Syndrom'].isin(top_syndroms)]
    
    # Total SNs for the main report date range
    total_sns_for_day = df['SerialNumber'].nunique()
    
    # Prepare report data: one row per SN
    report_rows = []
    for syndrom in top_syndroms:
        golden_img, defect_img, description = get_syndrom_db_info(syndrom)
        syndrom_fails = top_fail_df[top_fail_df['Syndrom'] == syndrom]
        for uut in syndrom_fails['UUT'].unique():
            uut_df = syndrom_fails[syndrom_fails['UUT'] == uut]
            for shift in ['1st Shift', '2nd Shift']:
                shift_fails = uut_df[uut_df['Shift'] == shift]
                fail_count = len(shift_fails)
                # Calculate unique SNs for this UUT and shift
                total_sns_for_shift = pd.Series(df[(df['UUT'] == uut) & (df['Shift'] == shift)]['SerialNumber']).nunique()
                rate = f"{(fail_count/total_sns_for_shift*100):.2f}%" if total_sns_for_shift > 0 else "N/A"
                for sn in shift_fails['SerialNumber'].astype(str):
                    report_rows.append({
                        'Monitor Name': syndrom,
                        'UUT': uut,
                        'Shift': shift,
                        'Rate': rate,
                        'SN': sn,
                        'Golden Image': golden_img,
                        'Defect Image': defect_img,
                        'Description': description or ''
                    })
    
    # Create DataFrame for Excel
    report_df = pd.DataFrame(report_rows)
    if not report_df.empty:
        report_df = report_df[['Monitor Name', 'UUT', 'Shift', 'Rate', 'SN', 'Golden Image', 'Defect Image', 'Description']]
        
        # Write to Excel with all columns
        with pd.ExcelWriter(REPORT_FILE, engine='openpyxl') as writer:
            report_df.to_excel(writer, index=False, sheet_name='Top 3 Syndroms')
        
        # Now add images using openpyxl
        wb = load_workbook(REPORT_FILE)
        ws = wb['Top 3 Syndroms']
        # Merge and insert images per unique syndrom
        create_merged_image_and_description_cells(ws, report_rows, syndrom_col=1, golden_img_col=6, defect_img_col=7, desc_col=8)
        # Merge cells for Monitor Name, UUT, Shift, and Rate
        for col_idx in [0, 1, 2, 3]:  # Monitor Name, UUT, Shift, Rate
            merge_consecutive_cells(ws, col_idx)
        
        # Ask if user wants trend charts
        print("\nDo you want to generate trend charts? (y/n): ", end="")
        trend_choice = input().strip().lower()
        
        if trend_choice == 'y':
            # Get date selection for trend analysis
            trend_start, trend_end = get_user_date_selection(file_dates, "trend analysis")
            if trend_start and trend_end:
                # Re-use the same pre-filtering idea for the trend window
                trend_files = [info['file'] for info in file_dates if not (info['max_date'] < trend_start or info['min_date'] > trend_end)]
                trend_df = load_data_duckdb(trend_start, trend_end, trend_files)
                if trend_df is not None:
                    # Calculate trend data using the same top 3 syndroms from main report
                    daily_df, weekly_df = calculate_trend_data(trend_df, top_syndroms)
                    
                    # Create trend charts
                    create_trend_charts(wb, daily_df, weekly_df, top_syndroms)
                    print("Trend charts added to Excel file!")
        
        wb.save(REPORT_FILE)
        print(f'\nReport generated: {REPORT_FILE}')
        print(f'Date range: {start_date} to {end_date}')

        # Load recipients and send email
        recipients = load_recipients()
        if recipients:
            # Create summary table for email (without SNs)
            summary_df = create_email_summary_table(report_rows)
            html_table = create_html_table(summary_df)
            
            # Initialize chart files list
            chart_files = []
            
            # Generate chart images if trend data exists
            if 'daily_df' in locals() and 'weekly_df' in locals():
                chart_files = generate_chart_images(daily_df, weekly_df, top_syndroms, start_date, end_date)
            
            # Send email with charts and table
            send_email_with_charts(recipients, chart_files, html_table, start_date, end_date)
        else:
            print("No recipients found in recipients.txt, skipping email.")

    else:
        print("No failed tests found for the selected date range!")

if __name__ == '__main__':
    main() 